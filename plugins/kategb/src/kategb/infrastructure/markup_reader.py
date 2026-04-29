from __future__ import annotations

from pathlib import Path
from typing import Any

from kategb.domain.models import CrystalInfo, LayerInfo


class OpenPyxlCrystalInfoReader:
    """Reads the legacy KateGB workbook layout without leaking openpyxl into the domain layer."""

    def read(self, path: Path) -> CrystalInfo:
        try:
            import openpyxl
        except ImportError as exc:
            raise RuntimeError("Для чтения Excel-разметки нужна зависимость 'openpyxl'.") from exc
        workbook = openpyxl.load_workbook(path, data_only=True)
        layers: dict[str, LayerInfo] = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            if getattr(sheet, "sheet_state", "visible") == "hidden":
                continue
            layer = self._read_layer(sheet_name, sheet)
            if layer is not None:
                layers[sheet_name] = layer
        if not layers:
            raise ValueError("В файле разметки не найдены листы слоев.")
        return CrystalInfo(layers=layers)

    def _read_layer(self, name: str, sheet: Any) -> LayerInfo | None:
        done_cell = sheet.cell(row=1, column=2).value
        if str(done_cell).lower() == "v":
            return LayerInfo(name=name, author_frames={}, frames_in_layer=0, frames_in_row=0, done=True)

        authors = self._read_authors(sheet)
        if not authors:
            return None
        author_frames: dict[str, list[int]] = {author: [] for author, _, _ in authors}
        mismatched: list[int] = []
        frames_in_layer = 0
        frames_in_row = 0

        row = 1
        while True:
            row_has_frames = False
            col = 2
            while True:
                cell = sheet.cell(row=row, column=col)
                if cell.value is None:
                    break
                try:
                    frame_number = int(cell.value)
                except (TypeError, ValueError):
                    break
                row_has_frames = True
                frames_in_layer += 1
                matched = False
                for author, color_index, tint in authors:
                    if cell.fill.start_color.index == color_index and cell.fill.start_color.tint == tint:
                        author_frames[author].append(frame_number)
                        matched = True
                        break
                if not matched:
                    mismatched.append(frame_number)
                col += 1
            if row_has_frames and frames_in_row == 0:
                frames_in_row = col - 2
            if not row_has_frames:
                break
            row += 1

        cif_folder, jpg_folder = self._read_folder_paths(sheet, row)
        return LayerInfo(
            name=name,
            author_frames={author: tuple(frames) for author, frames in author_frames.items()},
            frames_in_layer=frames_in_layer,
            frames_in_row=frames_in_row,
            cif_folder=Path(cif_folder) if cif_folder else None,
            jpg_folder=Path(jpg_folder) if jpg_folder else None,
            mismatched_cells=tuple(mismatched),
        )

    def _read_authors(self, sheet: Any) -> list[tuple[str, Any, float]]:
        authors: list[tuple[str, Any, float]] = []
        row = 1
        while True:
            name = sheet.cell(row=row, column=1).value
            marker = sheet.cell(row=row, column=2).value
            if name is None or marker is not None:
                break
            cell = sheet.cell(row=row, column=1)
            authors.append((str(name), cell.fill.start_color.index, cell.fill.start_color.tint))
            row += 1
        return authors

    def _read_folder_paths(self, sheet: Any, first_empty_frame_row: int) -> tuple[str | None, str | None]:
        row = first_empty_frame_row
        while row < first_empty_frame_row + 20:
            label = str(sheet.cell(row=row, column=1).value or "").lower()
            value = sheet.cell(row=row, column=2).value
            if value and ("cif" in label or not label):
                cif = str(value)
                jpg_value = sheet.cell(row=row + 1, column=2).value
                return cif, str(jpg_value) if jpg_value else None
            row += 1
        return None, None
