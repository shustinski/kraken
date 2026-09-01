"""Deterministic statistics with drill-down to the underlying activities.

The report adapter accepts both the public, stable activity names and the
CamelCase domain-event names emitted by the application layer.  Classification
of the generic ``ArtifactVersionCreated`` event is deliberately based on its
provenance/format payload: creating a version is not necessarily an import or a
vectorization.
"""

from __future__ import annotations

import csv
import json
from collections import Counter, defaultdict
from dataclasses import dataclass, field, replace
from datetime import UTC, datetime, time, timedelta, tzinfo
from enum import StrEnum
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping


@dataclass(frozen=True, slots=True)
class ActivityRecord:
    event_id: str
    recorded_at: datetime
    event_type: str
    project_id: str
    layer_id: str | None = None
    layer_type: str | None = None
    representation_id: str | None = None
    frame_id: str | None = None
    actor_id: str | None = None
    performer_id: str | None = None
    tool: str | None = None
    status: str | None = None
    bytes_count: int = 0
    payload: Mapping[str, Any] = field(default_factory=dict)

    def __post_init__(self) -> None:
        if self.recorded_at.tzinfo is None:
            raise ValueError("Activity timestamps must be timezone-aware")
        if isinstance(self.bytes_count, bool) or not isinstance(self.bytes_count, int):
            raise ValueError("Activity bytes_count must be an integer")

    @classmethod
    def from_event(cls, event: Any) -> ActivityRecord:
        """Build a report row from a domain ``EventEnvelope``.

        ``Any`` is intentional here: the reports extra is an infrastructure
        adapter and does not require callers to import the domain type merely
        to convert a stored/upcast event.
        """

        payload = dict(event.payload)
        representation_snapshot = payload.get("representation")
        if not isinstance(representation_snapshot, Mapping):
            representation_snapshot = {}
        program = getattr(event, "program", None)
        raw_size = payload.get("size_bytes", payload.get("bytes_count", 0))
        size = raw_size if isinstance(raw_size, int) and not isinstance(raw_size, bool) else 0
        return cls(
            event_id=str(event.event_id),
            recorded_at=event.recorded_at,
            event_type=str(event.event_type),
            project_id=str(event.project_id),
            layer_id=_optional_text(
                payload.get("layer_id") or representation_snapshot.get("layer_id")
            ),
            layer_type=_optional_text(payload.get("layer_type")),
            representation_id=_optional_text(
                payload.get("representation_id")
                or payload.get("target_representation_id")
                or representation_snapshot.get("id")
            ),
            frame_id=_optional_text(payload.get("frame_id")),
            actor_id=str(event.actor.principal_id),
            performer_id=_optional_text(getattr(event, "performer_id", None)),
            tool=(
                _optional_text(getattr(program, "name", None))
                or _optional_text(payload.get("tool_name"))
                or _optional_text(payload.get("tool"))
            ),
            status=_optional_text(payload.get("status")),
            bytes_count=size,
            payload=payload,
        )


@dataclass(frozen=True, slots=True)
class ReportFilters:
    start: datetime
    end: datetime
    project_ids: frozenset[str] = frozenset()
    layer_ids: frozenset[str] = frozenset()
    layer_types: frozenset[str] = frozenset()
    representation_ids: frozenset[str] = frozenset()
    performer_ids: frozenset[str] = frozenset()
    actor_ids: frozenset[str] = frozenset()
    tools: frozenset[str] = frozenset()
    event_types: frozenset[str] = frozenset()
    statuses: frozenset[str] = frozenset()

    def __post_init__(self) -> None:
        if self.start.tzinfo is None or self.end.tzinfo is None:
            raise ValueError("Report timestamps must be timezone-aware")
        if self.end < self.start:
            raise ValueError("Report end must not precede its start")

    def matches(self, item: ActivityRecord) -> bool:
        if not self.start <= item.recorded_at <= self.end:
            return False
        comparisons = (
            (self.project_ids, item.project_id),
            (self.layer_ids, item.layer_id),
            (self.layer_types, item.layer_type),
            (self.representation_ids, item.representation_id),
            (self.performer_ids, item.performer_id),
            (self.actor_ids, item.actor_id),
            (self.tools, item.tool),
            (self.event_types, item.event_type),
            (self.statuses, item.status),
        )
        return all(not accepted or value in accepted for accepted, value in comparisons)


@dataclass(frozen=True, slots=True)
class ReportMetrics:
    values: Mapping[str, int | float]
    by_project: Mapping[str, Mapping[str, int]]
    by_performer: Mapping[str, Mapping[str, int]]
    by_vector_layer: Mapping[str, Mapping[str, int]]
    event_ids_by_metric: Mapping[str, tuple[str, ...]]


class MetricValueKind(StrEnum):
    COUNT = "count"
    BYTES = "bytes"
    RATE = "rate"
    DURATION = "duration"


class MetricChartKind(StrEnum):
    BAR = "bar"
    LINE = "line"


class ReportGranularity(StrEnum):
    DAY = "day"
    WEEK = "week"
    MONTH = "month"
    YEAR = "year"


@dataclass(frozen=True, slots=True)
class MetricDefinition:
    key: str
    label: str
    description: str
    value_kind: MetricValueKind = MetricValueKind.COUNT
    chart_kind: MetricChartKind = MetricChartKind.BAR


@dataclass(frozen=True, slots=True)
class PresentedMetric:
    definition: MetricDefinition
    raw_value: int | float
    formatted_value: str


@dataclass(frozen=True, slots=True)
class ReportBucket:
    start: datetime
    end: datetime
    label: str
    metrics: ReportMetrics


@dataclass(frozen=True, slots=True)
class ReportSeries:
    granularity: ReportGranularity
    buckets: tuple[ReportBucket, ...]


METRIC_DEFINITIONS: tuple[MetricDefinition, ...] = (
    MetricDefinition("imported_files", "Импортировано файлов", "Количество импортированных файлов."),
    MetricDefinition(
        "imported_bytes",
        "Объём импортированных данных",
        "Суммарный размер импортированных файлов.",
        MetricValueKind.BYTES,
    ),
    MetricDefinition(
        "created_artifact_versions",
        "Создано версий файлов",
        "Количество созданных версий файлов и результатов обработки.",
    ),
    MetricDefinition(
        "binary_representations",
        "Создано бинарных представлений",
        "Количество созданных бинарных представлений изображений.",
    ),
    MetricDefinition(
        "vectorization_operations",
        "Операций векторизации",
        "Общее количество выполненных операций векторизации.",
    ),
    MetricDefinition(
        "unique_first_vectorized_frames",
        "Впервые векторизовано кадров",
        "Количество кадров, впервые векторизованных в выбранном периоде.",
    ),
    MetricDefinition("work_issued", "Выдано заданий", "Количество выданных заданий на проверку."),
    MetricDefinition(
        "returned_changed",
        "Возвращено с изменениями",
        "Количество результатов, возвращённых исполнителем с изменениями.",
    ),
    MetricDefinition(
        "returned_unchanged",
        "Возвращено без изменений",
        "Количество результатов, возвращённых без изменений.",
    ),
    MetricDefinition(
        "returned_missing",
        "Возвращено как отсутствующие",
        "Количество результатов, отмеченных исполнителем как отсутствующие.",
    ),
    MetricDefinition("accepted", "Принято", "Количество принятых результатов проверки."),
    MetricDefinition(
        "changes_requested",
        "Отправлено на доработку",
        "Количество результатов, для которых запрошена доработка.",
    ),
    MetricDefinition(
        "backlog",
        "Ожидают приёмки",
        "Разница между выданными и принятыми заданиями за выбранный период.",
    ),
    MetricDefinition(
        "rework_rate",
        "Доля доработок",
        "Доля запросов на доработку среди принятых и возвращённых на доработку результатов.",
        MetricValueKind.RATE,
        MetricChartKind.LINE,
    ),
    MetricDefinition(
        "average_turnaround_seconds",
        "Среднее время выполнения",
        "Среднее время от выдачи задания до возврата результата.",
        MetricValueKind.DURATION,
        MetricChartKind.LINE,
    ),
    MetricDefinition("overdue", "Просрочено", "Количество событий с просроченным статусом."),
    MetricDefinition(
        "incomplete_jobs",
        "Незавершённых заданий",
        "Количество частично выполненных заданий и заданий, требующих восстановления.",
    ),
    MetricDefinition("plugin_failures", "Ошибок плагинов", "Количество завершившихся ошибкой заданий плагинов."),
)

_METRIC_DEFINITIONS_BY_KEY = {definition.key: definition for definition in METRIC_DEFINITIONS}
_MONTH_NAMES = (
    "",
    "январь",
    "февраль",
    "март",
    "апрель",
    "май",
    "июнь",
    "июль",
    "август",
    "сентябрь",
    "октябрь",
    "ноябрь",
    "декабрь",
)


# Dotted values are the stable activity taxonomy. CamelCase values are the
# corresponding domain events already emitted (or reserved) by application use
# cases. Generic events receive additional payload-aware classification below.
_METRIC_EVENT_TYPES: dict[str, frozenset[str]] = {
    "created_artifact_versions": frozenset(
        {
            "ArtifactVersionCreated",
            "ArtifactVersionImported",
            "ArtifactImported",
            "artifact.version.created",
            "artifact.vector.created",
            "artifact.imported",
        }
    ),
    "vectorization_operations": frozenset(
        {"VectorizationCompleted", "FrameVectorized", "artifact.vector.created", "frame.vectorized"}
    ),
    "binary_representations": frozenset(
        {"BinaryRepresentationCreated", "representation.binary.created"}
    ),
    "imported_files": frozenset({"ArtifactImported", "ArtifactVersionImported", "artifact.imported"}),
    "work_issued": frozenset(
        {"WorkBatchIssued", "ReviewBatchIssued", "work_batch.issued", "review_batch.issued"}
    ),
    "returned_changed": frozenset({"ReviewReturnedChanged", "review.returned_changed"}),
    "returned_unchanged": frozenset({"ReviewReturnedUnchanged", "review.returned_unchanged"}),
    "returned_missing": frozenset({"ReviewReturnedMissing", "review.returned_missing"}),
    "accepted": frozenset({"ReviewAccepted", "ReviewCandidateAccepted", "review.accepted"}),
    "changes_requested": frozenset({"ReviewChangesRequested", "review.changes_requested"}),
    "plugin_failures": frozenset({"PluginJobFailed", "plugin_job.failed"}),
    "incomplete_jobs": frozenset(
        {"PluginJobPartial", "PluginJobRecoveryRequired", "plugin_job.partial", "plugin_job.recovery_required"}
    ),
}


def _optional_text(value: object) -> str | None:
    return None if value is None else str(value)


def _semantic_values(item: ActivityRecord) -> frozenset[str]:
    fields = (
        "activity_type",
        "operation",
        "origin",
        "source",
        "kind",
        "artifact_kind",
        "representation_kind",
        "pixel_format",
    )
    return frozenset(str(item.payload[key]).strip().casefold() for key in fields if item.payload.get(key) is not None)


def _is_vector_artifact(item: ActivityRecord) -> bool:
    semantic = _semantic_values(item)
    if semantic.intersection({"vector", "vectorized", "vectorization", "contour"}):
        return True
    media_type = str(item.payload.get("media_type", "")).casefold()
    filename = str(item.payload.get("filename", "")).casefold()
    return "cif" in media_type or filename.endswith(".cif")


def _is_binary_representation(item: ActivityRecord) -> bool:
    semantic = _semantic_values(item)
    if bool(item.payload.get("binary")):
        return True
    return bool(
        semantic.intersection(
            {"binary", "binary_image", "binary-image", "neuralimage", "1-bit", "0/255", "0,255"}
        )
    )


def _is_import(item: ActivityRecord) -> bool:
    semantic = _semantic_values(item)
    return any(value in {"import", "imported", "managed_copy", "external_link"} or value.startswith("import:") for value in semantic)


def _metric_categories(item: ActivityRecord) -> frozenset[str]:
    categories = {metric for metric, names in _METRIC_EVENT_TYPES.items() if item.event_type in names}
    if item.event_type in {"ArtifactVersionCreated", "artifact.version.created"}:
        if _is_vector_artifact(item):
            categories.add("vectorization_operations")
        if _is_import(item):
            categories.add("imported_files")
    if item.event_type in {"RepresentationCreated", "representation.created"} and _is_binary_representation(item):
        categories.add("binary_representations")
    return frozenset(categories)


class _MetricsAccumulator:
    """One-pass aggregate builder; only required drill-down IDs are retained."""

    def __init__(
        self,
        filters: ReportFilters,
        *,
        first_vectorization: Mapping[tuple[str, str, str], ActivityRecord] | None = None,
    ) -> None:
        self.filters = filters
        self.counts: Counter[str] = Counter()
        self.drill_down: dict[str, list[tuple[datetime, str]]] = defaultdict(list)
        self.first_vectorization = dict(first_vectorization or {})
        self.project_counters: dict[str, Counter[str]] = defaultdict(Counter)
        self.performer_counters: dict[str, Counter[str]] = defaultdict(Counter)
        self.vector_layer_counters: dict[str, Counter[str]] = defaultdict(Counter)
        self.imported_bytes = 0
        self.issued = 0
        self.accepted = 0
        self.rework = 0
        self.overdue = 0
        self.turnaround_total = 0.0
        self.turnaround_count = 0

    def add(self, item: ActivityRecord) -> bool:
        categories = _metric_categories(item)
        if (
            "vectorization_operations" in categories
            and item.frame_id is not None
            and item.layer_id is not None
        ):
            key = (
                item.project_id,
                item.representation_id or item.layer_id,
                item.frame_id,
            )
            existing = self.first_vectorization.get(key)
            if existing is None or (item.recorded_at, item.event_id) < (existing.recorded_at, existing.event_id):
                self.first_vectorization[key] = item

        if not self.filters.matches(item):
            return False

        for metric in categories:
            self.counts[metric] += 1
            self.drill_down[metric].append((item.recorded_at, item.event_id))
        if "imported_files" in categories:
            self.imported_bytes += max(0, item.bytes_count)
        self.issued += "work_issued" in categories
        self.accepted += "accepted" in categories
        self.rework += "changes_requested" in categories
        self.overdue += item.status == "overdue"
        if categories.intersection({"returned_changed", "returned_unchanged", "returned_missing"}):
            duration = item.payload.get("turnaround_seconds")
            if isinstance(duration, (int, float)) and not isinstance(duration, bool):
                self.turnaround_total += float(duration)
                self.turnaround_count += 1

        self.project_counters[item.project_id][item.event_type] += 1
        if item.performer_id:
            self.performer_counters[item.performer_id][item.event_type] += 1
        if item.representation_id:
            # Count every occurrence. A vector layer may pass the same workflow
            # stage repeatedly and each pass remains visible in statistics.
            self.vector_layer_counters[item.representation_id][item.event_type] += 1
        return True

    def finish(self) -> ReportMetrics:
        values: dict[str, int | float] = {
            metric: self.counts[metric] for metric in _METRIC_EVENT_TYPES
        }
        drill_down: dict[str, tuple[str, ...]] = {
            metric: tuple(event_id for _, event_id in sorted(self.drill_down[metric]))
            for metric in _METRIC_EVENT_TYPES
        }

        # "First" is determined over the complete supplied history, not merely
        # over the report period.  The earliest event is then tested against all
        # filters, including the selected period/performer/tool.
        first_in_scope = sorted(
            (item for item in self.first_vectorization.values() if self.filters.matches(item)),
            key=lambda item: (item.recorded_at, item.event_id),
        )
        values["unique_first_vectorized_frames"] = len(first_in_scope)
        drill_down["unique_first_vectorized_frames"] = tuple(item.event_id for item in first_in_scope)
        values["imported_bytes"] = self.imported_bytes
        values["backlog"] = max(0, self.issued - self.accepted)
        values["rework_rate"] = (
            0.0 if not self.accepted and not self.rework else self.rework / (self.accepted + self.rework)
        )
        values["overdue"] = self.overdue
        values["average_turnaround_seconds"] = (
            0.0 if not self.turnaround_count else self.turnaround_total / self.turnaround_count
        )
        return ReportMetrics(
            values=values,
            by_project={key: dict(value) for key, value in self.project_counters.items()},
            by_performer={key: dict(value) for key, value in self.performer_counters.items()},
            by_vector_layer={
                key: dict(value) for key, value in self.vector_layer_counters.items()
            },
            event_ids_by_metric=drill_down,
        )


def _safe_tabular_text(value: object) -> object:
    """Prevent spreadsheet formula interpretation in CSV and XLSX exports."""

    if not isinstance(value, str) or not value:
        return value
    stripped = value.lstrip(" \t\r\n")
    if value[0] in "\t\r\n" or (stripped and stripped[0] in "=+-@"):
        return "'" + value
    return value


def metric_definition(key: str) -> MetricDefinition:
    try:
        return _METRIC_DEFINITIONS_BY_KEY[key]
    except KeyError as exc:
        raise KeyError(f"Unknown report metric: {key}") from exc


def _format_decimal(value: float) -> str:
    rendered = f"{value:.1f}".rstrip("0").rstrip(".")
    return rendered.replace(".", ",")


def format_metric_value(definition: MetricDefinition, value: int | float) -> str:
    if definition.value_kind is MetricValueKind.COUNT:
        return f"{int(value):,}".replace(",", " ")
    if definition.value_kind is MetricValueKind.RATE:
        return f"{_format_decimal(float(value) * 100.0)} %"
    if definition.value_kind is MetricValueKind.BYTES:
        amount = max(0.0, float(value))
        units = ("Б", "КБ", "МБ", "ГБ", "ТБ")
        unit_index = 0
        while amount >= 1024.0 and unit_index < len(units) - 1:
            amount /= 1024.0
            unit_index += 1
        rendered = str(int(amount)) if unit_index == 0 else _format_decimal(amount)
        return f"{rendered} {units[unit_index]}"
    seconds = max(0, int(round(float(value))))
    hours, remainder = divmod(seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    parts = []
    if hours:
        parts.append(f"{hours} ч")
    if minutes:
        parts.append(f"{minutes} мин")
    if seconds or not parts:
        parts.append(f"{seconds} с")
    return " ".join(parts)


def present_metrics(values: Mapping[str, int | float]) -> tuple[PresentedMetric, ...]:
    return tuple(
        PresentedMetric(
            definition=definition,
            raw_value=values.get(definition.key, 0),
            formatted_value=format_metric_value(definition, values.get(definition.key, 0)),
        )
        for definition in METRIC_DEFINITIONS
    )


def _period_start(value: datetime, granularity: ReportGranularity) -> datetime:
    if granularity is ReportGranularity.DAY:
        period_date = value.date()
    elif granularity is ReportGranularity.WEEK:
        period_date = value.date() - timedelta(days=value.weekday())
    elif granularity is ReportGranularity.MONTH:
        period_date = value.date().replace(day=1)
    else:
        period_date = value.date().replace(month=1, day=1)
    return datetime.combine(period_date, time.min, value.tzinfo)


def _next_period(value: datetime, granularity: ReportGranularity) -> datetime:
    if granularity is ReportGranularity.DAY:
        return datetime.combine(value.date() + timedelta(days=1), time.min, value.tzinfo)
    if granularity is ReportGranularity.WEEK:
        return datetime.combine(value.date() + timedelta(days=7), time.min, value.tzinfo)
    if granularity is ReportGranularity.MONTH:
        year = value.year + (1 if value.month == 12 else 0)
        month = 1 if value.month == 12 else value.month + 1
        return datetime(year, month, 1, tzinfo=value.tzinfo)
    return datetime(value.year + 1, 1, 1, tzinfo=value.tzinfo)


def _period_label(start: datetime, end: datetime, granularity: ReportGranularity) -> str:
    if granularity is ReportGranularity.DAY:
        return start.strftime("%d.%m.%Y")
    if granularity is ReportGranularity.WEEK:
        last_day = (end - timedelta(microseconds=1)).date()
        return f"{start:%d.%m}–{last_day:%d.%m.%Y}"
    if granularity is ReportGranularity.MONTH:
        return f"{_MONTH_NAMES[start.month]} {start.year}"
    return str(start.year)


def _first_vectorizations(records: Iterable[ActivityRecord]) -> dict[tuple[str, str, str], ActivityRecord]:
    first: dict[tuple[str, str, str], ActivityRecord] = {}
    for item in records:
        if (
            "vectorization_operations" not in _metric_categories(item)
            or item.frame_id is None
            or item.layer_id is None
        ):
            continue
        key = (item.project_id, item.representation_id or item.layer_id, item.frame_id)
        existing = first.get(key)
        if existing is None or (item.recorded_at, item.event_id) < (existing.recorded_at, existing.event_id):
            first[key] = item
    return first


class ReportService:
    def iter_selected(self, records: Iterable[ActivityRecord], filters: ReportFilters) -> Iterator[ActivityRecord]:
        """Filter without buffering; callers supply canonical event order."""

        return (item for item in records if filters.matches(item))

    def select(self, records: Iterable[ActivityRecord], filters: ReportFilters) -> tuple[ActivityRecord, ...]:
        return tuple(sorted(self.iter_selected(records, filters), key=lambda item: (item.recorded_at, item.event_id)))

    def aggregate(self, records: Iterable[ActivityRecord], filters: ReportFilters) -> ReportMetrics:
        accumulator = _MetricsAccumulator(filters)
        for item in records:
            accumulator.add(item)
        return accumulator.finish()

    def aggregate_series(
        self,
        records: Iterable[ActivityRecord],
        filters: ReportFilters,
        granularity: ReportGranularity,
        *,
        timezone: tzinfo,
    ) -> ReportSeries:
        source = tuple(records)
        local_start = filters.start.astimezone(timezone)
        local_end = filters.end.astimezone(timezone)
        calendar_start = _period_start(local_start, granularity)
        seeded_first = _first_vectorizations(source)
        bucket_data: list[
            tuple[datetime, datetime, str, _MetricsAccumulator]
        ] = []
        accumulators_by_start: dict[datetime, _MetricsAccumulator] = {}

        current = calendar_start
        while current <= local_end:
            following = _next_period(current, granularity)
            bucket_start = max(filters.start, current.astimezone(UTC))
            bucket_end = min(filters.end, following.astimezone(UTC) - timedelta(microseconds=1))
            if bucket_start <= bucket_end:
                accumulator = _MetricsAccumulator(
                    replace(filters, start=bucket_start, end=bucket_end),
                    first_vectorization=seeded_first,
                )
                bucket_data.append(
                    (
                        bucket_start,
                        bucket_end,
                        _period_label(current, following, granularity),
                        accumulator,
                    )
                )
                accumulators_by_start[current] = accumulator
            current = following

        for item in source:
            if not filters.start <= item.recorded_at <= filters.end:
                continue
            key = _period_start(item.recorded_at.astimezone(timezone), granularity)
            selected_accumulator = accumulators_by_start.get(key)
            if selected_accumulator is not None:
                selected_accumulator.add(item)

        return ReportSeries(
            granularity=granularity,
            buckets=tuple(
                ReportBucket(start, end, label, accumulator.finish())
                for start, end, label, accumulator in bucket_data
            ),
        )

    def write_csv(
        self,
        destination: Path | str,
        records: Iterable[ActivityRecord],
        filters: ReportFilters,
        *,
        assume_sorted: bool = False,
    ) -> Path:
        """Write an UTF-8 journal.

        ``assume_sorted=True`` enables constant-memory streaming for a cursor
        already ordered by ``(recorded_at, event_id)``.  The default preserves
        the historical deterministic sorting contract for arbitrary iterables.
        """

        output = Path(destination)
        output.parent.mkdir(parents=True, exist_ok=True)
        selected: Iterable[ActivityRecord]
        selected = self.iter_selected(records, filters) if assume_sorted else self.select(records, filters)
        with output.open("w", encoding="utf-8", newline="") as stream:
            writer = csv.writer(stream, lineterminator="\n")
            writer.writerow(
                (
                    "event_id",
                    "recorded_at",
                    "event_type",
                    "project_id",
                    "layer_id",
                    "layer_type",
                    "representation_id",
                    "frame_id",
                    "actor_id",
                    "performer_id",
                    "tool",
                    "status",
                    "bytes_count",
                    "payload_json",
                )
            )
            for item in selected:
                row = (
                    item.event_id,
                    item.recorded_at.isoformat(),
                    item.event_type,
                    item.project_id,
                    item.layer_id or "",
                    item.layer_type or "",
                    item.representation_id or "",
                    item.frame_id or "",
                    item.actor_id or "",
                    item.performer_id or "",
                    item.tool or "",
                    item.status or "",
                    item.bytes_count,
                    json.dumps(dict(item.payload), ensure_ascii=False, sort_keys=True, separators=(",", ":")),
                )
                writer.writerow(tuple(_safe_tabular_text(value) for value in row))
        return output

    def write_xlsx(
        self,
        destination: Path | str,
        records: Iterable[ActivityRecord],
        filters: ReportFilters,
        *,
        assume_sorted: bool = False,
        include_series: bool = False,
        timezone: tzinfo | None = None,
    ) -> Path:
        try:
            from openpyxl import Workbook
        except ImportError as exc:  # pragma: no cover - optional dependency
            raise RuntimeError("Install Kraken with the 'reports' extra to export XLSX") from exc

        # write_only avoids retaining worksheet cells.  For an ordered database
        # cursor, assume_sorted also avoids retaining report rows themselves.
        series_records: tuple[ActivityRecord, ...] = ()
        source: Iterable[ActivityRecord]
        if include_series:
            series_records = tuple(records)
            source = (
                series_records
                if assume_sorted
                else sorted(series_records, key=lambda item: (item.recorded_at, item.event_id))
            )
        elif assume_sorted:
            source = records
        else:
            source = sorted(records, key=lambda item: (item.recorded_at, item.event_id))
        workbook = Workbook(write_only=True)
        summary = workbook.create_sheet("Сводка")
        series_sheets = {
            ReportGranularity.DAY: workbook.create_sheet("По дням"),
            ReportGranularity.WEEK: workbook.create_sheet("По неделям"),
            ReportGranularity.MONTH: workbook.create_sheet("По месяцам"),
            ReportGranularity.YEAR: workbook.create_sheet("По годам"),
        } if include_series else {}
        projects = workbook.create_sheet("Projects")
        layers_sheet = workbook.create_sheet("Layers")
        vector_layers_sheet = workbook.create_sheet("Vector layers")
        performers = workbook.create_sheet("Performers")
        tools_sheet = workbook.create_sheet("Tools")
        events = workbook.create_sheet("Events")
        events.append(
            (
                "Event ID",
                "Recorded at",
                "Type",
                "Project",
                "Layer",
                "Vector layer",
                "Frame",
                "Actor",
                "Performer",
                "Tool",
                "Status",
            )
        )
        accumulator = _MetricsAccumulator(filters)
        layer_counters: dict[str, Counter[str]] = defaultdict(Counter)
        tool_counters: dict[str, Counter[str]] = defaultdict(Counter)

        for item in source:
            selected = accumulator.add(item)
            if not selected:
                continue
            if item.layer_id:
                layer_counters[item.layer_id][item.event_type] += 1
            if item.tool:
                tool_counters[item.tool][item.event_type] += 1
            events.append(
                tuple(
                    _safe_tabular_text(value)
                    for value in (
                        item.event_id,
                        item.recorded_at.isoformat(),
                        item.event_type,
                        item.project_id,
                        item.layer_id,
                        item.representation_id,
                        item.frame_id,
                        item.actor_id,
                        item.performer_id,
                        item.tool,
                        item.status,
                    )
                )
            )

        metrics = accumulator.finish()
        summary.append(("Показатель", "Значение"))
        for metric in present_metrics(metrics.values):
            summary.append(
                (
                    _safe_tabular_text(metric.definition.label),
                    _safe_tabular_text(metric.formatted_value),
                )
            )
        local_timezone = timezone or datetime.now().astimezone().tzinfo or UTC
        for granularity, sheet in series_sheets.items():
            series = self.aggregate_series(
                series_records,
                filters,
                granularity,
                timezone=local_timezone,
            )
            sheet.append(("Период", *(definition.label for definition in METRIC_DEFINITIONS)))
            for bucket in series.buckets:
                presented = present_metrics(bucket.metrics.values)
                sheet.append(
                    (
                        bucket.label,
                        *(metric.formatted_value for metric in presented),
                    )
                )
        self._write_counter_sheet(projects, "Project", metrics.by_project)
        self._write_counter_sheet(layers_sheet, "Layer", layer_counters)
        self._write_counter_sheet(
            vector_layers_sheet, "Vector layer", metrics.by_vector_layer
        )
        self._write_counter_sheet(performers, "Performer", metrics.by_performer)
        self._write_counter_sheet(tools_sheet, "Tool", tool_counters)

        output = Path(destination)
        output.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output)
        return output

    @staticmethod
    def _write_counter_sheet(sheet: Any, item_heading: str, counters: Mapping[str, Mapping[str, int]]) -> None:
        event_types = sorted({event for counter in counters.values() for event in counter})
        sheet.append(tuple(_safe_tabular_text(value) for value in (item_heading, *event_types)))
        for key, counter in sorted(counters.items()):
            sheet.append(
                (_safe_tabular_text(key), *(counter.get(event, 0) for event in event_types))
            )

    def write_pdf(
        self,
        destination: Path | str,
        records: Iterable[ActivityRecord],
        filters: ReportFilters,
        *,
        cyrillic_font: Path | str,
    ) -> Path:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.pdfgen.canvas import Canvas
        except ImportError as exc:  # pragma: no cover - optional dependency
            raise RuntimeError("Install Kraken with the 'reports' extra to export PDF") from exc
        font = Path(cyrillic_font).resolve(strict=True)
        metrics = self.aggregate(records, filters)
        output = Path(destination)
        output.parent.mkdir(parents=True, exist_ok=True)
        font_name = "KrakenCyrillic"
        pdfmetrics.registerFont(TTFont(font_name, str(font)))
        canvas = Canvas(str(output), pagesize=A4)
        _, height = A4
        canvas.setFont(font_name, 15)
        canvas.drawString(40, height - 45, "Kraken — отчёт по активности")
        canvas.setFont(font_name, 9)
        canvas.drawString(40, height - 65, f"Период: {filters.start.isoformat()} — {filters.end.isoformat()}")
        y = height - 90
        for key, value in sorted(metrics.values.items()):
            canvas.drawString(45, y, f"{key}: {value}")
            y -= 14
            if y < 45:
                canvas.showPage()
                canvas.setFont(font_name, 9)
                y = height - 45
        canvas.save()
        return output


__all__ = ["ActivityRecord", "ReportFilters", "ReportMetrics", "ReportService"]
