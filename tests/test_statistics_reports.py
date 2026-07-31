from __future__ import annotations

import tempfile
import unittest
from importlib.util import find_spec
from datetime import UTC, datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from kraken_manager.infrastructure.reports import (
    ActivityRecord,
    ReportFilters,
    ReportGranularity,
    ReportService,
    format_metric_value,
    metric_definition,
)


class StatisticsReportTests(unittest.TestCase):
    def test_human_readable_metric_values(self) -> None:
        self.assertEqual(
            "1 234",
            format_metric_value(metric_definition("imported_files"), 1234),
        )
        self.assertEqual(
            "1,5 МБ",
            format_metric_value(metric_definition("imported_bytes"), 1572864),
        )
        self.assertEqual(
            "12,5 %",
            format_metric_value(metric_definition("rework_rate"), 0.125),
        )
        self.assertEqual(
            "1 ч 2 мин 3 с",
            format_metric_value(metric_definition("average_turnaround_seconds"), 3723),
        )

    def test_calendar_series_use_local_boundaries_and_project_filter(self) -> None:
        timezone = ZoneInfo("Europe/Moscow")

        def at(year: int, month: int, day: int, hour: int, minute: int = 0) -> datetime:
            return datetime(year, month, day, hour, minute, tzinfo=timezone).astimezone(UTC)

        records = (
            ActivityRecord("jan", at(2026, 1, 31, 23, 59), "artifact.imported", "p"),
            ActivityRecord("boundary", at(2026, 2, 1, 0, 0), "artifact.imported", "p"),
            ActivityRecord("monday", at(2026, 2, 2, 10, 0), "artifact.imported", "p"),
            ActivityRecord("other", at(2026, 2, 1, 12, 0), "artifact.imported", "other"),
        )
        filters = ReportFilters(
            at(2026, 1, 31, 12, 0),
            at(2026, 2, 2, 12, 0),
            project_ids=frozenset({"p"}),
        )
        reports = ReportService()

        daily = reports.aggregate_series(
            records,
            filters,
            ReportGranularity.DAY,
            timezone=timezone,
        )
        weekly = reports.aggregate_series(
            records,
            filters,
            ReportGranularity.WEEK,
            timezone=timezone,
        )
        monthly = reports.aggregate_series(
            records,
            filters,
            ReportGranularity.MONTH,
            timezone=timezone,
        )

        self.assertEqual(
            ("31.01.2026", "01.02.2026", "02.02.2026"),
            tuple(bucket.label for bucket in daily.buckets),
        )
        self.assertEqual(
            (1, 1, 1),
            tuple(bucket.metrics.values["imported_files"] for bucket in daily.buckets),
        )
        self.assertEqual(
            (2, 1),
            tuple(bucket.metrics.values["imported_files"] for bucket in weekly.buckets),
        )
        self.assertEqual(
            (1, 2),
            tuple(bucket.metrics.values["imported_files"] for bucket in monthly.buckets),
        )

    def test_series_first_vectorization_uses_complete_history(self) -> None:
        timezone = ZoneInfo("Europe/Moscow")
        old = datetime(2026, 1, 1, 10, tzinfo=timezone).astimezone(UTC)
        current = datetime(2026, 2, 1, 10, tzinfo=timezone).astimezone(UTC)
        records = (
            ActivityRecord("old", old, "frame.vectorized", "p", "l", frame_id="repeated"),
            ActivityRecord("repeat", current, "frame.vectorized", "p", "l", frame_id="repeated"),
            ActivityRecord("first", current, "frame.vectorized", "p", "l", frame_id="new"),
        )
        filters = ReportFilters(current - timedelta(hours=1), current + timedelta(hours=1))

        series = ReportService().aggregate_series(
            records,
            filters,
            ReportGranularity.DAY,
            timezone=timezone,
        )

        self.assertEqual(1, series.buckets[0].metrics.values["unique_first_vectorized_frames"])
        self.assertEqual(
            ("first",),
            series.buckets[0].metrics.event_ids_by_metric["unique_first_vectorized_frames"],
        )

    def test_metrics_distinguish_unique_frames_and_operations(self) -> None:
        now = datetime.now(UTC)
        records = (
            ActivityRecord("e1", now, "frame.vectorized", "p", "l", frame_id="f", performer_id="worker"),
            ActivityRecord("e2", now, "artifact.vector.created", "p", "l", frame_id="f", performer_id="worker"),
            ActivityRecord("e3", now, "artifact.imported", "p", "l", frame_id="f", bytes_count=123),
            ActivityRecord("e4", now, "review.changes_requested", "p", "l", frame_id="f"),
        )
        filters = ReportFilters(now - timedelta(seconds=1), now + timedelta(seconds=1))
        metrics = ReportService().aggregate(records, filters)
        self.assertEqual(2, metrics.values["vectorization_operations"])
        self.assertEqual(1, metrics.values["unique_first_vectorized_frames"])
        self.assertEqual(123, metrics.values["imported_bytes"])
        self.assertEqual(("e1",), metrics.event_ids_by_metric["unique_first_vectorized_frames"])

    def test_application_event_names_are_classified_from_payload(self) -> None:
        now = datetime.now(UTC)
        records = (
            ActivityRecord(
                "version",
                now,
                "ArtifactVersionCreated",
                "p",
                "l",
                frame_id="f",
                bytes_count=50,
                payload={"filename": "1_1.cif", "source": "import"},
            ),
            ActivityRecord(
                "representation",
                now,
                "RepresentationCreated",
                "p",
                "l",
                payload={"kind": "image", "source": "NeuralImage", "binary": True},
            ),
        )
        filters = ReportFilters(now - timedelta(seconds=1), now + timedelta(seconds=1))

        metrics = ReportService().aggregate(records, filters)

        self.assertEqual(1, metrics.values["created_artifact_versions"])
        self.assertEqual(1, metrics.values["vectorization_operations"])
        self.assertEqual(1, metrics.values["imported_files"])
        self.assertEqual(50, metrics.values["imported_bytes"])
        self.assertEqual(1, metrics.values["binary_representations"])

    def test_first_vectorization_is_computed_over_complete_supplied_history(self) -> None:
        now = datetime.now(UTC)
        records = (
            ActivityRecord("old", now - timedelta(days=2), "FrameVectorized", "p", "l", frame_id="f1"),
            ActivityRecord("repeat", now, "FrameVectorized", "p", "l", frame_id="f1"),
            ActivityRecord("first", now, "FrameVectorized", "p", "l", frame_id="f2"),
        )
        filters = ReportFilters(now - timedelta(hours=1), now + timedelta(hours=1))

        metrics = ReportService().aggregate(iter(records), filters)

        self.assertEqual(2, metrics.values["vectorization_operations"])
        self.assertEqual(1, metrics.values["unique_first_vectorized_frames"])
        self.assertEqual(("first",), metrics.event_ids_by_metric["unique_first_vectorized_frames"])

    def test_repeated_stages_are_counted_per_vector_layer(self) -> None:
        now = datetime.now(UTC)
        records = (
            ActivityRecord(
                "one",
                now,
                "frame.vectorized",
                "p",
                "l",
                representation_id="vectors-a",
                frame_id="f",
            ),
            ActivityRecord(
                "two",
                now,
                "frame.vectorized",
                "p",
                "l",
                representation_id="vectors-a",
                frame_id="f",
            ),
        )
        filters = ReportFilters(now - timedelta(seconds=1), now + timedelta(seconds=1))

        metrics = ReportService().aggregate(records, filters)

        self.assertEqual(
            {"frame.vectorized": 2},
            metrics.by_vector_layer["vectors-a"],
        )

    def test_csv_is_utf8_normalized_event_journal(self) -> None:
        now = datetime.now(UTC)
        record = ActivityRecord("событие", now, "artifact.imported", "проект", payload={"заметка": "тест"})
        with tempfile.TemporaryDirectory() as temporary:
            output = ReportService().write_csv(
                Path(temporary) / "report.csv",
                (record,),
                ReportFilters(now - timedelta(seconds=1), now + timedelta(seconds=1)),
            )
            contents = output.read_text(encoding="utf-8")
            self.assertIn("событие", contents)
            self.assertIn("заметка", contents)

    def test_csv_neutralizes_spreadsheet_formulas(self) -> None:
        now = datetime.now(UTC)
        record = ActivityRecord(
            "=HYPERLINK(\"https://invalid\")",
            now,
            "artifact.imported",
            "+project",
            actor_id="  @actor",
            tool="-tool",
        )
        with tempfile.TemporaryDirectory() as temporary:
            output = ReportService().write_csv(
                Path(temporary) / "report.csv",
                (record,),
                ReportFilters(now - timedelta(seconds=1), now + timedelta(seconds=1)),
            )
            contents = output.read_text(encoding="utf-8")

        self.assertIn("'=HYPERLINK", contents)
        self.assertIn("'+project", contents)
        self.assertIn("'  @actor", contents)
        self.assertIn("'-tool", contents)

    def test_csv_can_stream_an_already_sorted_cursor(self) -> None:
        now = datetime.now(UTC)
        consumed: list[str] = []

        def cursor():
            for event_id in ("one", "two"):
                consumed.append(event_id)
                yield ActivityRecord(event_id, now, "ProjectCreated", "p")

        with tempfile.TemporaryDirectory() as temporary:
            ReportService().write_csv(
                Path(temporary) / "report.csv",
                cursor(),
                ReportFilters(now - timedelta(seconds=1), now + timedelta(seconds=1)),
                assume_sorted=True,
            )

        self.assertEqual(["one", "two"], consumed)

    @unittest.skipUnless(find_spec("openpyxl"), "openpyxl is an optional reports dependency")
    def test_xlsx_neutralizes_spreadsheet_formulas(self) -> None:
        from openpyxl import load_workbook

        now = datetime.now(UTC)
        record = ActivityRecord("=1+1", now, "ProjectCreated", "+project", tool="@tool")
        with tempfile.TemporaryDirectory() as temporary:
            output = ReportService().write_xlsx(
                Path(temporary) / "report.xlsx",
                (record,),
                ReportFilters(now - timedelta(seconds=1), now + timedelta(seconds=1)),
            )
            workbook = load_workbook(output, read_only=True, data_only=False)
            events = workbook["Events"]
            row = next(events.iter_rows(min_row=2, values_only=False))
            values = tuple(cell.value for cell in row)
            data_types = tuple(cell.data_type for cell in row)
            workbook.close()

        self.assertEqual("'=1+1", values[0])
        self.assertNotEqual("f", data_types[0])
        self.assertEqual("'+project", values[3])
        self.assertEqual("'@tool", values[9])

    @unittest.skipUnless(find_spec("openpyxl"), "openpyxl is an optional reports dependency")
    def test_xlsx_contains_human_readable_summary_and_series(self) -> None:
        from openpyxl import load_workbook

        now = datetime(2026, 7, 30, 10, tzinfo=UTC)
        record = ActivityRecord("import", now, "artifact.imported", "p", bytes_count=1536)
        filters = ReportFilters(now - timedelta(days=1), now + timedelta(days=1))
        with tempfile.TemporaryDirectory() as temporary:
            output = ReportService().write_xlsx(
                Path(temporary) / "report.xlsx",
                (record,),
                filters,
                include_series=True,
                timezone=UTC,
            )
            workbook = load_workbook(output, read_only=True, data_only=True)
            self.assertEqual(
                ["Сводка", "По дням", "По неделям", "По месяцам", "По годам"],
                workbook.sheetnames[:5],
            )
            summary_rows = tuple(workbook["Сводка"].iter_rows(values_only=True))
            daily_rows = tuple(workbook["По дням"].iter_rows(values_only=True))
            workbook.close()

        self.assertEqual(("Показатель", "Значение"), summary_rows[0])
        self.assertEqual(("Импортировано файлов", "1"), summary_rows[1])
        self.assertEqual("Период", daily_rows[0][0])
        self.assertEqual("Импортировано файлов", daily_rows[0][1])


if __name__ == "__main__":
    unittest.main()
